from __future__ import annotations

import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlencode, urljoin

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import (
    Browser,
    BrowserContext,
    Page,
    TimeoutError as PlaywrightTimeoutError,
    sync_playwright,
)


# ============================================================
# KONFIGURASI
# ============================================================

BASE_URL = "https://maganghub.kemnaker.go.id"
LIST_PATH = "/magang-nasional/lowongan"

OUTPUT_DIR = Path(__file__).resolve().parent / "hasil_scraping"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

HEADLESS = True
REQUEST_DELAY_SECONDS = 1.5
PAGE_TIMEOUT_MS = 120_000
MAX_LOAD_MORE_ATTEMPTS = 20

FILTER_PARAMS: dict[str, str] = {
    "keyword": "",
    "education_level[0][id]": "bachelor",
    "education_level[0][label]": "Sarjana",
    "city_id[0][id]": "49e68da2-eaa7-401c-b4e4-0feaf679287f",
    "city_id[0][label]": "Kota Pekanbaru",
    "study_program[0][id]": "246e6e8a-413d-4dff-9d82-161b6222f8d0",
    "study_program[0][label]": "Ilmu Komputer",
    "study_program[1][id]": "8bdce5f1-553c-40b4-a5f5-9e5ff9387b60",
    "study_program[1][label]": "Sistem Informasi",
    "study_program[2][id]": "ae73f44b-835b-498b-a10a-95eaa32fbe0b",
    "study_program[2][label]": "Sistem Informasi",
    "study_program[3][id]": "03b570a2-c678-4413-aa47-bfed8308fac5",
    "study_program[3][label]": "Teknik Informatika",
    "sort": "most_applicants",
}

LIST_URL = f"{BASE_URL}{LIST_PATH}?{urlencode(FILTER_PARAMS)}"


@dataclass
class CardData:
    url: str
    card_text: str | None = None
    title_card: str | None = None
    organizer_card: str | None = None
    study_program_card: str | None = None
    location_card: str | None = None
    education_card: str | None = None
    workdays_card: int | None = None
    quota_card: int | None = None
    applicants_card: int | None = None
    holidays_card: str | None = None


# ============================================================
# FUNGSI UTILITAS
# ============================================================

def clean_text(value: Any) -> str | None:
    """Membersihkan spasi, tab, dan baris kosong berlebihan."""
    if value is None:
        return None

    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def extract_int(text: str | None, pattern: str) -> int | None:
    """Mengambil bilangan bulat pertama berdasarkan regex."""
    if not text:
        return None

    match = re.search(pattern, text, flags=re.IGNORECASE)
    if not match:
        return None

    try:
        return int(match.group(1))
    except (TypeError, ValueError):
        return None


def extract_section(
    text: str | None,
    start_heading: str,
    end_headings: list[str],
) -> str | None:
    """Mengambil teks di antara sebuah judul dan judul berikutnya."""
    if not text:
        return None

    lower_text = text.lower()
    start_position = lower_text.find(start_heading.lower())

    if start_position == -1:
        return None

    content_start = start_position + len(start_heading)
    remaining = text[content_start:]
    remaining_lower = remaining.lower()
    content_end = len(remaining)

    for heading in end_headings:
        position = remaining_lower.find(heading.lower())
        if 0 <= position < content_end:
            content_end = position

    return clean_text(remaining[:content_end])


def first_matching_text(page: Page, selectors: list[str]) -> str | None:
    """Mengambil teks pertama yang tidak kosong dari daftar selector."""
    for selector in selectors:
        locator = page.locator(selector)

        try:
            count = locator.count()
        except Exception:
            continue

        for index in range(min(count, 5)):
            try:
                text = clean_text(locator.nth(index).inner_text(timeout=3_000))
            except Exception:
                continue

            if text:
                return text

    return None


def extract_location(text: str | None) -> str | None:
    if not text:
        return None

    match = re.search(
        r"\b(Kota|Kabupaten)\s+[A-Za-zÀ-ÿ0-9 .,'’()/-]+?"
        r"(?=\s+(?:Diploma|Sarjana|Profesi|\d+\s*hari|Kuota|Pelamar|Hari Libur|$))",
        text,
        flags=re.IGNORECASE,
    )
    return clean_text(match.group(0)) if match else None


def extract_education(text: str | None) -> str | None:
    if not text:
        return None

    found = [
        level
        for level in ("Diploma", "Sarjana", "Profesi")
        if re.search(rf"\b{re.escape(level)}\b", text, flags=re.IGNORECASE)
    ]

    return ", ".join(dict.fromkeys(found)) if found else None


def competition_category(ratio: float | None) -> str | None:
    if ratio is None or pd.isna(ratio):
        return None
    if ratio <= 5:
        return "Relatif rendah"
    if ratio <= 10:
        return "Sedang"
    if ratio <= 20:
        return "Tinggi"
    return "Sangat tinggi"


def title_from_slug(url: str) -> str:
    slug = url.rstrip("/").split("/")[-1]
    slug = re.sub(
        r"-[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$",
        "",
        slug,
        flags=re.IGNORECASE,
    )
    return slug.replace("-", " ").title()


# ============================================================
# PARSING KARTU DAFTAR
# ============================================================

def parse_card_text(url: str, card_text: str | None) -> CardData:
    text = clean_text(card_text) or ""

    quota = extract_int(text, r"Kuota\s*:?\s*(\d+)")
    applicants = extract_int(text, r"Pelamar\s*:?\s*(\d+)")
    workdays = extract_int(text, r"(\d+)\s*hari\s*/?\s*minggu")

    holidays = None
    holiday_match = re.search(
        r"Hari Libur\s+(.+?)(?=\s+(?:Kuota|Pelamar|$))",
        text,
        flags=re.IGNORECASE,
    )
    if holiday_match:
        holidays = clean_text(holiday_match.group(1))

    title = None
    if text:
        title_part = re.split(
            r"\s+(?=(?:Kota|Kabupaten)\s+Pekanbaru\b)",
            text,
            maxsplit=1,
            flags=re.IGNORECASE,
        )[0]
        title = clean_text(title_part)

    return CardData(
        url=url,
        card_text=text or None,
        title_card=title,
        location_card=extract_location(text),
        education_card=extract_education(text),
        workdays_card=workdays,
        quota_card=quota,
        applicants_card=applicants,
        holidays_card=holidays,
    )


# ============================================================
# SCRAPING HALAMAN DAFTAR
# ============================================================

def get_total_listings(page_text: str) -> int | None:
    return extract_int(page_text, r"Ditemukan\s+(\d+)\s+lowongan")


def click_load_more(page: Page) -> bool:
    """Menekan tombol muat tambahan jika ditemukan."""
    button_patterns = [
        re.compile(r"Muat\s+Lebih\s+Banyak", re.IGNORECASE),
        re.compile(r"Lihat\s+Lebih\s+Banyak", re.IGNORECASE),
        re.compile(r"Tampilkan\s+Lebih\s+Banyak", re.IGNORECASE),
        re.compile(r"Selanjutnya", re.IGNORECASE),
    ]

    for pattern in button_patterns:
        locator = page.get_by_role("button", name=pattern)

        if locator.count() == 0:
            continue

        button = locator.first

        try:
            if button.is_visible() and button.is_enabled():
                button.scroll_into_view_if_needed()
                button.click(timeout=5_000)
                return True
        except Exception:
            continue

    return False


def collect_listing_cards(page: Page) -> list[CardData]:
    print("=" * 75)
    print("MEMBUKA HALAMAN LOWONGAN")
    print("=" * 75)
    print(LIST_URL)
    print()

    page.goto(
        LIST_URL,
        wait_until="domcontentloaded",
        timeout=PAGE_TIMEOUT_MS,
    )
    page.wait_for_timeout(5_000)

    page_text = clean_text(page.locator("body").inner_text()) or ""
    total_expected = get_total_listings(page_text)

    if total_expected is not None:
        print(f"Jumlah lowongan menurut website: {total_expected}")
    else:
        print("Jumlah total lowongan tidak dapat dibaca.")

    detail_selector = 'a[href*="/magang-nasional/lowongan/"]'
    previous_unique_count = -1

    for _ in range(MAX_LOAD_MORE_ATTEMPTS):
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        page.wait_for_timeout(1_250)

        hrefs = page.locator(detail_selector).evaluate_all(
            "elements => elements.map(element => element.href)"
        )
        unique_detail_urls = {
            href.split("#")[0]
            for href in hrefs
            if re.search(
                r"/magang-nasional/lowongan/[^/?#]+/?$",
                href.split("?")[0],
                flags=re.IGNORECASE,
            )
        }

        current_unique_count = len(unique_detail_urls)

        if total_expected is not None and current_unique_count >= total_expected:
            break

        button_clicked = click_load_more(page)
        if button_clicked:
            page.wait_for_timeout(2_000)
            previous_unique_count = current_unique_count
            continue

        if current_unique_count == previous_unique_count:
            break

        previous_unique_count = current_unique_count

    links = page.locator(detail_selector)
    cards: list[CardData] = []
    visited_urls: set[str] = set()

    for index in range(links.count()):
        link = links.nth(index)

        href = link.get_attribute("href")
        if not href:
            continue

        absolute_url = urljoin(BASE_URL, href).split("#")[0]
        path_without_query = absolute_url.split("?")[0].rstrip("/")

        if not re.search(
            r"/magang-nasional/lowongan/[^/?#]+$",
            path_without_query,
            flags=re.IGNORECASE,
        ):
            continue

        if absolute_url in visited_urls:
            continue

        try:
            card_text = link.inner_text(timeout=3_000)
        except Exception:
            try:
                card_text = link.locator("xpath=..").inner_text(timeout=3_000)
            except Exception:
                card_text = None

        cards.append(parse_card_text(absolute_url, card_text))
        visited_urls.add(absolute_url)

    print(f"URL detail unik yang ditemukan: {len(cards)}")
    print()

    if total_expected and len(cards) < total_expected:
        print(
            "PERINGATAN: jumlah URL yang ditemukan lebih sedikit daripada "
            "jumlah yang ditampilkan website."
        )
        print(
            "Script tetap dilanjutkan. Kemungkinan tombol pemuatan tambahan "
            "atau struktur halaman berubah."
        )
        print()

    return cards


# ============================================================
# SCRAPING HALAMAN DETAIL
# ============================================================

def scrape_detail(page: Page, card: CardData) -> dict[str, Any]:
    page.goto(
        card.url,
        wait_until="domcontentloaded",
        timeout=PAGE_TIMEOUT_MS,
    )
    page.wait_for_timeout(1_500)

    body_raw = page.locator("body").inner_text(timeout=20_000)
    body_text = clean_text(body_raw) or ""

    title = first_matching_text(
        page,
        [
            "main h1",
            "main h2",
            "h1",
            "h2",
        ],
    )

    if not title or title.lower() in {"lowongan magang", "maganghub"}:
        title = card.title_card or title_from_slug(card.url)

    quota = extract_int(body_text, r"Kuota\s*:?\s*(\d+)")
    applicants = extract_int(body_text, r"Pelamar\s*:?\s*(\d+)")
    workdays = extract_int(
        body_text,
        r"(\d+)\s*hari(?:\s*kerja)?\s*(?:per|/)\s*minggu",
    )

    if quota is None:
        quota = card.quota_card
    if applicants is None:
        applicants = card.applicants_card
    if workdays is None:
        workdays = card.workdays_card

    description = extract_section(
        body_text,
        "Deskripsi Lowongan",
        [
            "Kualifikasi",
            "Program studi",
            "Tingkat pendidikan",
            "Skill yang Bakal Kamu Dapat",
            "Hari Libur",
            "Lokasi Magang",
            "Durasi Magang",
        ],
    )

    qualification = extract_section(
        body_text,
        "Kualifikasi",
        [
            "Program studi",
            "Tingkat pendidikan",
            "Skill yang Bakal Kamu Dapat",
            "Hari Libur",
            "Lokasi Magang",
            "Durasi Magang",
        ],
    )

    study_program = extract_section(
        body_text,
        "Program studi",
        [
            "Tingkat pendidikan",
            "Hari kerja",
            "Skill yang Bakal Kamu Dapat",
            "Deskripsi Lowongan",
            "Lokasi Magang",
        ],
    )

    education_section = extract_section(
        body_text,
        "Tingkat pendidikan",
        [
            "Program studi",
            "Hari kerja",
            "Skill yang Bakal Kamu Dapat",
            "Deskripsi Lowongan",
            "Lokasi Magang",
        ],
    )

    education = extract_education(education_section) or card.education_card

    skills = extract_section(
        body_text,
        "Skill yang Bakal Kamu Dapat",
        [
            "Hari Libur",
            "Lokasi Magang",
            "Durasi Magang",
            "Kuota",
            "Pelamar",
        ],
    )

    holidays = extract_section(
        body_text,
        "Hari Libur",
        [
            "Lokasi Magang",
            "Durasi Magang",
            "Kuota",
            "Pelamar",
            "Lamar Sekarang",
        ],
    )
    holidays = holidays or card.holidays_card

    location_detail = extract_section(
        body_text,
        "Lokasi Magang",
        [
            "Durasi Magang",
            "Kuota",
            "Pelamar",
            "Lamar Sekarang",
        ],
    )

    location = extract_location(body_text) or card.location_card

    organizer = extract_section(
        body_text,
        "Penyelenggara",
        [
            "Deskripsi Lowongan",
            "Kualifikasi",
            "Program studi",
            "Tingkat pendidikan",
            "Lokasi Magang",
        ],
    )

    return {
        "judul_posisi": title,
        "penyelenggara": organizer,
        "lokasi": location,
        "lokasi_detail": location_detail,
        "tingkat_pendidikan": education,
        "program_studi": study_program,
        "hari_kerja_per_minggu": workdays,
        "hari_libur": holidays,
        "kuota": quota,
        "jumlah_pelamar": applicants,
        "deskripsi": description,
        "kualifikasi": qualification,
        "skill_yang_didapat": skills,
        "url": card.url,
        "status_scraping": "Berhasil",
        "error": None,
    }


def failed_result(card: CardData, error: Exception) -> dict[str, Any]:
    return {
        "judul_posisi": card.title_card or title_from_slug(card.url),
        "penyelenggara": card.organizer_card,
        "lokasi": card.location_card,
        "lokasi_detail": None,
        "tingkat_pendidikan": card.education_card,
        "program_studi": card.study_program_card,
        "hari_kerja_per_minggu": card.workdays_card,
        "hari_libur": card.holidays_card,
        "kuota": card.quota_card,
        "jumlah_pelamar": card.applicants_card,
        "deskripsi": None,
        "kualifikasi": None,
        "skill_yang_didapat": None,
        "url": card.url,
        "status_scraping": "Gagal",
        "error": f"{type(error).__name__}: {error}",
    }


# ============================================================
# PENGOLAHAN DATA
# ============================================================

def build_dataframe(rows: list[dict[str, Any]]) -> pd.DataFrame:
    df = pd.DataFrame(rows)

    if df.empty:
        return df

    df = df.drop_duplicates(subset=["url"]).reset_index(drop=True)

    for column in ("kuota", "jumlah_pelamar", "hari_kerja_per_minggu"):
        df[column] = pd.to_numeric(df[column], errors="coerce").astype("Int64")

    quota_numeric = pd.to_numeric(df["kuota"], errors="coerce")
    applicants_numeric = pd.to_numeric(df["jumlah_pelamar"], errors="coerce")

    valid_ratio = quota_numeric.gt(0) & applicants_numeric.notna()
    df["rasio_pelamar_per_kuota"] = pd.NA
    df.loc[valid_ratio, "rasio_pelamar_per_kuota"] = (
        applicants_numeric[valid_ratio] / quota_numeric[valid_ratio]
    ).round(2)
    df["rasio_pelamar_per_kuota"] = pd.to_numeric(
        df["rasio_pelamar_per_kuota"], errors="coerce"
    )

    valid_probability = quota_numeric.gt(0) & applicants_numeric.gt(0)
    df["estimasi_peluang_sederhana_persen"] = pd.NA
    df.loc[valid_probability, "estimasi_peluang_sederhana_persen"] = (
        quota_numeric[valid_probability]
        / applicants_numeric[valid_probability]
        * 100
    ).clip(upper=100).round(2)
    df["estimasi_peluang_sederhana_persen"] = pd.to_numeric(
        df["estimasi_peluang_sederhana_persen"], errors="coerce"
    )

    df["kategori_persaingan"] = df[
        "rasio_pelamar_per_kuota"
    ].apply(competition_category)

    df["tanggal_scraping"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    df = df.sort_values(
        by=["rasio_pelamar_per_kuota", "jumlah_pelamar"],
        ascending=[True, True],
        na_position="last",
    ).reset_index(drop=True)

    df.insert(0, "peringkat_peluang", range(1, len(df) + 1))
    return df


def format_excel(excel_path: Path) -> None:
    """Memberi format dasar pada workbook hasil."""
    from openpyxl import load_workbook

    workbook = load_workbook(excel_path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for column_index, column_cells in enumerate(
            worksheet.iter_cols(1, worksheet.max_column),
            start=1,
        ):
            max_length = 0

            for cell in column_cells:
                if cell.value is None:
                    continue

                value_length = len(str(cell.value))
                max_length = max(max_length, value_length)

                if cell.row > 1:
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True,
                    )

            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width = min(max(max_length + 2, 12), 55)

    workbook.save(excel_path)


def save_outputs(df: pd.DataFrame) -> tuple[Path, Path]:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = OUTPUT_DIR / f"lowongan_maganghub_pekanbaru_{timestamp}.xlsx"
    csv_path = OUTPUT_DIR / f"lowongan_maganghub_pekanbaru_{timestamp}.csv"

    recommendation_columns = [
        "peringkat_peluang",
        "judul_posisi",
        "penyelenggara",
        "lokasi",
        "kuota",
        "jumlah_pelamar",
        "rasio_pelamar_per_kuota",
        "estimasi_peluang_sederhana_persen",
        "kategori_persaingan",
        "program_studi",
        "tingkat_pendidikan",
        "url",
    ]

    filter_info = pd.DataFrame(
        {
            "Filter": [
                "Jenjang",
                "Lokasi",
                "Program Studi 1",
                "Program Studi 2",
                "Program Studi 3",
                "Program Studi 4",
                "Urutan",
                "URL",
            ],
            "Nilai": [
                "Sarjana",
                "Kota Pekanbaru",
                "Ilmu Komputer",
                "Sistem Informasi",
                "Sistem Informasi",
                "Teknik Informatika",
                "Pelamar terbanyak",
                LIST_URL,
            ],
        }
    )

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Semua Lowongan", index=False)
        df[recommendation_columns].to_excel(
            writer,
            sheet_name="Peringkat Peluang",
            index=False,
        )
        filter_info.to_excel(
            writer,
            sheet_name="Informasi Filter",
            index=False,
        )

    format_excel(excel_path)
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    return excel_path, csv_path


# ============================================================
# PROGRAM UTAMA
# ============================================================

def create_browser_context(browser: Browser) -> BrowserContext:
    return browser.new_context(
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/150.0.0.0 Safari/537.36"
        ),
        locale="id-ID",
        viewport={"width": 1440, "height": 1000},
        ignore_https_errors=False,
    )


def run_scraper() -> pd.DataFrame:
    rows: list[dict[str, Any]] = []

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            headless=HEADLESS,
            args=[
                "--disable-dev-shm-usage",
                "--disable-blink-features=AutomationControlled",
            ],
        )

        context = create_browser_context(browser)
        list_page = context.new_page()
        list_page.set_default_timeout(30_000)

        try:
            cards = collect_listing_cards(list_page)
        finally:
            list_page.close()

        if not cards:
            context.close()
            browser.close()
            raise RuntimeError(
                "Tidak ada URL lowongan yang ditemukan. "
                "Periksa koneksi internet atau struktur halaman website."
            )

        detail_page = context.new_page()
        detail_page.set_default_timeout(30_000)

        total = len(cards)

        for index, card in enumerate(cards, start=1):
            print(f"[{index:02d}/{total:02d}] {card.url}")

            try:
                result = scrape_detail(detail_page, card)
                rows.append(result)
                print("  Status: berhasil")
            except PlaywrightTimeoutError as error:
                rows.append(failed_result(card, error))
                print("  Status: gagal karena timeout")
            except Exception as error:
                rows.append(failed_result(card, error))
                print(f"  Status: gagal — {type(error).__name__}: {error}")

            time.sleep(REQUEST_DELAY_SECONDS)

        detail_page.close()
        context.close()
        browser.close()

    return build_dataframe(rows)


def print_summary(df: pd.DataFrame, excel_path: Path, csv_path: Path) -> None:
    success_count = int((df["status_scraping"] == "Berhasil").sum())
    failed_count = int((df["status_scraping"] == "Gagal").sum())

    print()
    print("=" * 75)
    print("SCRAPING SELESAI")
    print("=" * 75)
    print(f"Jumlah data : {len(df)}")
    print(f"Berhasil    : {success_count}")
    print(f"Gagal       : {failed_count}")
    print(f"Excel       : {excel_path}")
    print(f"CSV         : {csv_path}")
    print("=" * 75)

    preview_columns = [
        "peringkat_peluang",
        "judul_posisi",
        "kuota",
        "jumlah_pelamar",
        "rasio_pelamar_per_kuota",
        "kategori_persaingan",
    ]

    print()
    print("10 LOWONGAN DENGAN RASIO PERSAINGAN TERENDAH")
    print("-" * 75)
    print(df[preview_columns].head(10).to_string(index=False))


def main() -> int:
    try:
        df = run_scraper()

        if df.empty:
            print("Tidak ada data yang berhasil diproses.")
            return 1

        excel_path, csv_path = save_outputs(df)
        print_summary(df, excel_path, csv_path)
        return 0

    except KeyboardInterrupt:
        print("\nProgram dihentikan oleh pengguna.")
        return 130

    except ModuleNotFoundError as error:
        print("\nLibrary belum terpasang:")
        print(error)
        print("\nJalankan:")
        print("python -m pip install playwright pandas openpyxl")
        print("python -m playwright install chromium")
        return 1

    except Exception as error:
        print("\nProgram mengalami kesalahan:")
        print(f"{type(error).__name__}: {error}")
        return 1


if __name__ == "__main__":
    sys.exit(main())